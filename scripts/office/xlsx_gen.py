#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表格生成器 — 接收结构化 JSON 描述，用 openpyxl 渲染 .xlsx 文件。

用法:
  python xlsx_gen.py <input.json> <output_dir>

input.json 结构:
{
  "title": "工作簿标题",
  "sheets": [
    {
      "name": "Sheet 名称（可选）",
      "headers": ["列1", "列2", "列3"],
      "rows": [["a1","a2","a3"], ["b1","b2","b3"]],
      "note": "表底部备注（可选）"
    }
  ]
}

输出: 打印 JSON 到 stdout:
  {"success": true, "path": "D:/DesktopPetData/Documents/xxx.xlsx", "title": "..."}
"""

import json
import sys
import os
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRIMARY_HEX = "1F4E79"
ACCENT_HEX = "2E86C1"
LIGHT_HEX = "DE EBF7".replace(" ", "")
FONT_CN = "微软雅黑"

HEADER_FILL = PatternFill(start_color=PRIMARY_HEX, end_color=PRIMARY_HEX, fill_type="solid")
HEADER_FONT = Font(name=FONT_CN, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_CN, size=11)
NOTE_FONT = Font(name=FONT_CN, size=10, color="666666")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def sanitize_filename(name):
    return "".join(c for c in name if c not in '<>:"/\\|?*').strip() or "workbook"


def write_sheet(ws, sheet_desc):
    headers = sheet_desc.get("headers") or []
    rows = sheet_desc.get("rows") or []
    note = sheet_desc.get("note") or ""

    # 表头
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=str(h))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # 数据行
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    # 备注
    if note:
        nr = len(rows) + 3
        ws.cell(row=nr, column=1, value=note).font = NOTE_FONT

    # 列宽
    for ci in range(1, max(len(headers), 1) + 1):
        max_len = len(str(headers[ci - 1])) if headers else 8
        for row in rows[:20]:
            if ci <= len(row):
                try:
                    max_len = max(max_len, len(str(row[ci - 1])))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 1.8, 12), 40)

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"success": False, "error": "用法: python xlsx_gen.py <input.json> <output_dir>"}, ensure_ascii=False))
        return 1

    input_path = sys.argv[1]
    out_dir = sys.argv[2]

    try:
        # utf-8-sig 兼容 PowerShell 写入的 BOM，也兼容 Node 无 BOM 写入
        with open(input_path, "r", encoding="utf-8-sig") as f:
            desc = json.load(f)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"读取描述文件失败: {e}"}, ensure_ascii=False))
        return 1

    try:
        title = str(desc.get("title") or "工作簿")
        sheets_desc = desc.get("sheets") or []

        wb = Workbook()
        wb.remove(wb.active)  # 移除默认 sheet

        for i, sheet_desc in enumerate(sheets_desc):
            name = str(sheet_desc.get("name") or f"Sheet{i + 1}")
            ws = wb.create_sheet(title=name[:31])
            write_sheet(ws, sheet_desc)

        if not sheets_desc:
            ws = wb.create_sheet(title="Sheet1")
            write_sheet(ws, {"headers": [], "rows": []})

        os.makedirs(out_dir, exist_ok=True)
        safe_title = sanitize_filename(title)
        filename = f"{safe_title}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = os.path.join(out_dir, filename)
        wb.save(path)

        print(json.dumps({
            "success": True,
            "path": path.replace("\\", "/"),
            "title": title,
            "sheets": len(sheets_desc)
        }, ensure_ascii=False))
        return 0
    except Exception as e:
        print(json.dumps({"success": False, "error": f"生成失败: {e}"}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    sys.exit(main())
